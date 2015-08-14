#-*- coding: UTF-8 -*-

import sys
import time
import urllib
import urllib2
import requests
import numpy as np
from bs4 import BeautifulSoup
from openpyxl import Workbook

reload(sys)
sys.setdefaultencoding('utf8')


def movie_spider(movie_tag):
    page_num=0;
    count=1
    movie_list=[]
    try_times=0
    
    #Some User Agents
    hds=[{'User-Agent':'Mozilla/5.0 (Windows; U; Windows NT 6.1; en-US; rv:1.9.1.6) Gecko/20091201 Firefox/3.5.6'},\
         {'User-Agent':'Mozilla/5.0 (Windows NT 6.2) AppleWebKit/535.11 (KHTML, like Gecko) Chrome/17.0.963.12 Safari/535.11'},\
         {'User-Agent': 'Mozilla/5.0 (compatible; MSIE 10.0; Windows NT 6.2; Trident/6.0)'}]
         
    while(1):
        url="http://www.douban.com/tag/"+urllib.quote(movie_tag)+"/movie?start="+str(page_num*15)
        time.sleep(np.random.rand()*2)
        
        #Last Version
        try:
            req = urllib2.Request(url, headers=hds[page_num%len(hds)])
            source_code = urllib2.urlopen(req).read()
            plain_text=str(source_code)   
        except (urllib2.HTTPError, urllib2.URLError), e:
            print e
            continue
  
        ##Previous Version, IP is easy to be Forbidden
        #source_code = requests.get(url) 
        #plain_text = source_code.text  
        
        soup = BeautifulSoup(plain_text)
        list_soup = soup.find('div', {'class': 'mod movie-list'})
        
        try_times+=1;
        if list_soup==None and try_times<200:
            continue
        elif list_soup==None or len(list_soup)<=1:
            break # Break when no informatoin got after 200 times requesting
        
        for movie_info in list_soup.findAll('dd'):
            title = movie_info.find('a', {'class':'title'}).string.strip()
            desc = movie_info.find('div', {'class':'desc'}).string.strip()            
            year = filter(lambda x: x.isdigit(), desc) #找出字符串中的数字，即年份
		    desc_list = desc.split('year') #以年份为分隔符，将字符串分割
            
		    try:
                type = '类型： ' + '/'.join(desc_list[0].split('/')[:-1]) #找出影片的类型
            except:
                type = '类型： 暂无'
            try:
                director = '导演： ' + '/'.join(desc_list[-1].split('/')[1]) #找出导演
            except:
                director = '导演： 暂无'
		    try:
			    actor = '演员： ' + '/'.join(desc_list[-1].split('/')[2:]) # 找出演员
		    except:
			    actor = '演员： 暂无'
            try:
                rating = movie_info.find('span', {'class':'rating_nums'}).string.strip()
            except:
                rating='0.0'
            try:
                people_num = book_info.findAll('span')[2].string.strip()
                people_num=people_num.strip('人评价')
            except:
                people_num='0'
            
            movie_list.append([title,rating,people_num,type,year,director,actor])
            try_times=0 #set 0 when got valid information
        page_num+=1
        print "Downloading Information From Page %d" % page_num
    return movie_list


def do_spider(movie_tag_lists):
    movie_lists=[]
    for movie_tag in movie_tag_lists:
        movie_list=movie_spider(movie_tag)
        movie_list=sorted(movie_list,key=lambda x:x[1],reverse=True)
        movie_lists.append(movie_list)
    return movie_lists


def print_movie_lists_excel(movie_lists,movie_tag_lists):
    wb=Workbook(optimized_write=True)
    ws=[]
    for i in range(len(movie_tag_lists)):
        ws.append(wb.create_sheet(title=movie_tag_lists[i].decode())) #utf8->unicode
    for i in range(len(movie_tag_lists)): 
        ws[i].append(['序号','电影名','评分','评价人数','类型','年份','导演','演员'])
        count=1
        for bl in movie_lists[i]:
            ws[i].append([count,bl[0],float(bl[1]),int(bl[2]),bl[3],int(bl[4]),bl[5],bl[6]])
            count+=1
    save_path='movie_list'
    for i in range(len(movie_tag_lists)):
        save_path+=('-'+movie_tag_lists[i].decode())
    save_path+='.xlsx'
    wb.save(save_path)



if __name__=='__main__':
    #book_tag_lists = ['心理','判断与决策','算法','数据结构','经济','历史']
    #book_tag_lists = ['传记','哲学','编程','创业','理财','社会学','佛教']
    #book_tag_lists=['思想','科技','科学','web','股票','爱情','两性']
    movie_tag_lists=['经典','动画']
    movie_lists=do_spider(movie_tag_lists)
    print_movie_lists_excel(movie_lists,movie_tag_lists)