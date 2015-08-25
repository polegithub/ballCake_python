#-*- coding: UTF-8 -*-

import sys
import time
import datetime
import urllib
import urllib2
import requests
import numpy as np
import re
from bs4 import BeautifulSoup
from openpyxl import Workbook

from crawler.getBasicData.base import *
from crawler.directory.directoryManager import contentClass

reload(sys)
sys.setdefaultencoding('utf8')



class DoubanMovie(Base):
    def __init__(self, session):
        super(DoubanMovie, self).__init__(session)

        self.type = 2


    def startSearchMovie(self):
        #movie_tag_lists = ['爱情','喜剧','动画','科幻','剧情','动作']
        #movie_tag_lists = ['经典','悬疑','青春','犯罪','惊悚','文艺','纪录片']
        #movie_tag_lists=['励志','搞笑','恐怖','战争','短片','魔幻','传记']

        movie_tag_lists = ['搞笑']
        print('start search list:%s' % (movie_tag_lists[0]))

        movie_lists = self.do_spider(movie_tag_lists)
        self.print_movie_lists_excel(movie_lists,movie_tag_lists)

    def do_spider(self,movie_tag_lists):
        movie_lists=[]
        for movie_tag in movie_tag_lists:
            movie_list = self.movie_spider(movie_tag)
            if movie_list:
                movie_list = sorted(movie_list,key=lambda x:x[1],reverse=True)
                movie_lists.append(movie_list)
        return movie_lists


    def movie_spider(self,movie_tag):
        page_num=0
        count=1
        movie_list=[]
        try_times=0

        #Some User Agents
        hds=[{'User-Agent':'Mozilla/5.0 (Windows; U; Windows NT 6.1; en-US; rv:1.9.1.6) Gecko/20091201 Firefox/3.5.6'},\
             {'User-Agent':'Mozilla/5.0 (Windows NT 6.2) AppleWebKit/535.11 (KHTML, like Gecko) Chrome/17.0.963.12 Safari/535.11'},\
             {'User-Agent': 'Mozilla/5.0 (compatible; MSIE 10.0; Windows NT 6.2; Trident/6.0)'}]

        print('start movie_spider')

        while(page_num < 2):
            url="http://www.douban.com/tag/"+urllib.quote(movie_tag)+"/movie?start="+str(page_num*15)
            print('start get url',url)

            # #Last Version
            # try:
            #     req = urllib2.Request(url,heads = hds[page_num%len(hds)])
            #     source_code = urllib2.urlopen(req).read()
            #     plain_text=str(source_code)
            # except (urllib2.HTTPError, urllib2.URLError), e:
            #     print e
            #     continue

            request = urllib2.Request(url,headers=hds[page_num%len(hds)])
            soup = self.getSoupWithUrl(request)

            ##Previous Version, IP is easy to be Forbidden
            #source_code = requests.get(url)
            #plain_text = source_code.text

            list_soup = soup.find('div', {'class': 'mod movie-list'})

            try_times+=1
            if list_soup==None and try_times<200:
                continue
            elif list_soup==None or len(list_soup)<=1:
                break # Break when no informatoin got after 200 times requesting

            temp = 1

            for movie_info in list_soup.findAll('dd'):
                titleFull =movie_info.find('a', {'class':'title'})

                title = titleFull.string.strip()

                url = titleFull.get('href')

                print("It's time to get detail:",url)
                request = urllib2.Request(url)
                soup_detail = self.getSoupWithUrl(request)

                # if temp ==1:

                    # print('soup detail:',soup_detail)

                # temp+=1


                movie_model = self.getDetailModel(title,url,soup_detail)

                movie_list.append(movie_model)
                try_times=0 #set 0 when got valid information

                #chenglong
                break;

            page_num += 1
            print "Downloading Information From Page %d" % page_num
        return movie_list


    #打印到excel
    def print_movie_lists_excel(self,movie_lists,movie_tag_lists):
        """

        :rtype : object
        """
        wb=Workbook(optimized_write=True)
        print('print_movie_lists_excel')
        ws=[]
        for i in range(len(movie_tag_lists)):
            ws.append(wb.create_sheet(title=movie_tag_lists[i].decode())) #utf8->unicode
        for i in range(len(movie_tag_lists)):
            ws[i].append(['序号','电影名','评分','评价人数','类型','年份','导演','演员','五星','四星','三星','二星','一星','推荐列表'])
            count = 1
            for bl in movie_lists[i]:
                print('bl.len:',len(bl))

                # ws[i].append([count,bl[0],float(bl[1]),int(bl[2]),bl[3],(bl[4]),bl[5],bl[6],bl[7],bl[8],bl[9],bl[10],bl[11]])

                count+=1


            # 获取 存储路径
        f =  contentClass()
        save_path= f.getContentForDouBan(2)
        print('save path:',save_path)


        for i in range(len(movie_tag_lists)):
            save_path+=('-'+movie_tag_lists[i].decode())
        save_path+='.xlsx'
        wb.save(save_path)



    # 获取电影的详细信息
    def getDetailModel(self,title,url,soup_detail):
        try:
            director = soup_detail.find("a",attrs = {"rel":"v:directedBy"}).string.strip()
        except:
            director = "暂无"

        try:
            actors = soup_detail.findAll("a",attrs = {"rel":"v:starring"})
            actor_list = []
            for i in range(0,len(actors)):
                actor = actors[i].string.strip()
                actor_list.append(actor)
            actor = '/'.join(actor_list)
        except:
            actor = "暂无"

        try:
            type_lists = soup_detail.findAll("span",attrs = {"property":"v:genre"})
            type_list = []
            for i in range(0,len(type_lists)):
                type = type_lists[i].string.strip()
                type_list.append(type)
            type = '/'.join(type_list)
        except:
            type = "暂无"

        try:
            ReleaseDate_list = soup_detail.find("span",attrs = {"property":"v:initialReleaseDate"})
            ReleaseDate = ReleaseDate_list.string.strip()
        except:
            ReleaseDate = None

        try:
            rating_num = soup_detail.find("strong",attrs = {"property":"v:average"}).string.strip()
        except:
            rating_num = 0


        try:
            vote_num = soup_detail.find("span",attrs = {"property":"v:votes"}).string.strip()
        except:
            vote_num = 0

        dateAndRegion = ''
        region =''

        if ReleaseDate:
            dateAndRegion = ReleaseDate.split('(')
            if dateAndRegion:
                dateFormat =datetime.datetime.strptime(dateAndRegion[0],'%Y-%m-%d')
                dateInterval = time.mktime(dateFormat.timetuple())

                region = dateAndRegion[1].split(')')[0]

        movieId = self.getMovieIdFromUrl(url)

        basicModel= [movieId,title,rating_num,vote_num,type,dateFormat,region,director,url]
        self.insertMovieBasicModel(basicModel)

        percent = re.compile("\S\d%")
        stars_percent = soup_detail.find("div", attrs = {"class":"rating_wrap clearbox"}).findAll(text=percent)
        stars5_percent = stars_percent[0].strip()
        stars4_percent = stars_percent[1].strip()
        stars3_percent = stars_percent[2].strip()
        stars2_percent = stars_percent[3].strip()
        stars1_percent = stars_percent[4].strip()

        starsModel =[movieId,title,rating_num,vote_num,
                     stars5_percent,stars4_percent,stars3_percent,stars2_percent,stars1_percent]
        self.insertMovieScoreModel(starsModel)


        #soup 使用findall之后不可再使用findall（因为findall的结果是resultset,不再是soup了）
        recommendations = soup_detail.select("dd > a")

        # recommendationsIngfo = soup_detail.find("div", attrs = {"class":"recommendations-bd"})
        # recommendations = recommendationsIngfo.find("dd")
        recommendations_MovieInfo_list = []
        # if recommendations:
        #      findAllList =recommendations.findAll("a")
        #
        # if findAllList:
        if recommendations:
            index =1

            for movie_info_simple in recommendations:
                movieTitle = movie_info_simple.string.strip()
                movieUrl = movie_info_simple.get('href')
                movieIdNew = self.getMovieIdFromUrl(movieUrl)
                movieInfo_list = [movieTitle, movieUrl, movieIdNew]
                recommendations_MovieInfo_list.append(movieInfo_list)

                self.insertMovieSimilarModel(movieId,movieIdNew,index)

                index+=1


        movie_model =[title,rating_num,vote_num,type,ReleaseDate,director,actor,stars5_percent,stars4_percent,stars3_percent,stars2_percent,stars1_percent, recommendations_MovieInfo_list]

        return movie_model

    #比较通用的解析url的函数
    def getSoupWithUrl(self,req):
        time.sleep(np.random.rand()*2)

        try:
            source_code_detail = urllib2.urlopen(req).read()
            plain_text_detail=str(source_code_detail)
        except (urllib2.HTTPError, urllib2.URLError), e:
            print e

        if plain_text_detail:
            soup = BeautifulSoup(plain_text_detail)
        else:
            soup = None

        return soup

    # get movie through url
    def getMovieIdFromUrl(self,url):
        movieId = re.match(r'\d.*(?=/)',url)
        movieId = filter(lambda x:x.isdigit(), url)
        return  movieId


    #insert to mysql
    def insertMovieBasicModel(self,basicModel):

        # store data to mysql
        #    def __init__(self,movieId,movieName,score,scoredNum,type,year,firstRegion,director,url):
        self.insert_Moview_basic(movieId=basicModel[0],movieName=basicModel[1],score=basicModel[2],
                                 scoredNum=basicModel[3],type=basicModel[4],
                                 year=basicModel[5],firstRegion= basicModel[6],
                                 director = basicModel[7],url=basicModel[8])



    def insertMovieScoreModel(self,model):
        print ('score:',model,)
        self.insert_Movie_Score(movieId=model[0],movieName=model[1],totalScore=model[2],totalNum=model[3],
                                FiveScore=model[4],FourScore=model[5],ThreeScore=model[6],TwoScore=model[7],
                                OneScore=model[8])

    def insertMovieTagModel(self,movieId,movieName,tag,index,tagType):

        self.insert_Movie_tag(movieId=movieId,movieName=movieName,tag=tag,index=index,tagType=tagType)


    def insertMovieSimilarModel(self,movieId,recommendId,index):
        self.insert_Movie_recommendMovieId(movieId=movieId,moviewNewId=recommendId,index = index)





