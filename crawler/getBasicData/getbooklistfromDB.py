#-*- coding: UTF-8 -*-

import sys
import time
import urllib
import urllib2
import requests
import numpy as np
from bs4 import BeautifulSoup
from openpyxl import Workbook

from crawler.directory.directoryManager import contentClass
from  analyse.dbHandle import do_init_db

reload(sys)
sys.setdefaultencoding('utf8')

BOOk_TYPE = 1

def startSearchBoook():
    #book_tag_lists = ['心理','判断与决策','算法','数据结构','经济','历史']
    #book_tag_lists = ['传记','哲学','编程','创业','理财','社会学','佛教']
    #book_tag_lists=['思想','科技','科学','web','股票','爱情','两性']
    # book_tag_lists=['计算机','机器学习','linux','android','数据库','互联网']
    book_tag_lists=['计算机']

    book_lists=do_spider(book_tag_lists)
    print_book_lists_excel(book_lists,book_tag_lists)


def book_spider(book_tag):
    page_num=0
    count=1
    book_list=[]
    try_times=0
    
    #Some User Agents
    hds=[{'User-Agent':'Mozilla/5.0 (Windows; U; Windows NT 6.1; en-US; rv:1.9.1.6) Gecko/20091201 Firefox/3.5.6'},\
         {'User-Agent':'Mozilla/5.0 (Windows NT 6.2) AppleWebKit/535.11 (KHTML, like Gecko) Chrome/17.0.963.12 Safari/535.11'},\
         {'User-Agent': 'Mozilla/5.0 (compatible; MSIE 10.0; Windows NT 6.2; Trident/6.0)'}]
         
    while(page_num < 2):
        url="http://www.douban.com/tag/"+urllib.quote(book_tag)+"/book?start="+str(page_num*15)
        time.sleep(np.random.rand()*2)
        
        #Last Version
        try:
            req = urllib2.Request(url, headers=hds[page_num%len(hds)])
            source_code = urllib2.urlopen(req).read()
            plain_text=str(source_code)   
        except (urllib2.HTTPError, urllib2.URLError), e:
            print e
            continue
  
        ##Previous Version, IP is easy to be Forbidden
        #source_code = requests.get(url) 
        #plain_text = source_code.text  
        
        soup = BeautifulSoup(plain_text)
        list_soup = soup.find('div', {'class': 'mod book-list'})
        
        try_times+=1;
        if list_soup==None and try_times<200:
            continue
        elif list_soup==None or len(list_soup)<=1:
            break # Break when no informatoin got after 200 times requesting
        
        for book_info in list_soup.findAll('dd'):
            title = book_info.find('a', {'class':'title'}).string.strip()
            desc = book_info.find('div', {'class':'desc'}).string.strip()
            desc_list = desc.split('/')
            
            try:
                author_info = '作者/译者： ' + '/'.join(desc_list[0:-3])
            except:
                author_info='作者/译者： 暂无'
            try:
                pub_info = '出版信息： ' + '/'.join(desc_list[-3:])
            except:
                pub_info='出版信息： 暂无'
            try:
                rating = book_info.find('span', {'class':'rating_nums'}).string.strip()
            except:
                rating='0.0'
            try:
                people_num = book_info.findAll('span')[2].string.strip()
                people_num=people_num.strip('人评价')
            except:
                people_num='0'
            
            book_list.append([title,rating,people_num,author_info,pub_info])
            try_times=0 #set 0 when got valid information
        page_num+=1
        print "Downloading Information From Page %d" % page_num
    print('get single book list')
    return book_list


def do_spider(book_tag_lists):
    book_lists=[]
    for book_tag in book_tag_lists:
        book_list=book_spider(book_tag)
        book_list=sorted(book_list,key=lambda x:x[1],reverse=True)
        book_lists.append(book_list)

    print('get book lists')
    return book_lists


def print_book_lists_excel(book_lists,book_tag_lists):
    wb=Workbook(optimized_write=True)

    print('start print_book_lists_excel')

    create_db()

    ws=[]
    for i in range(len(book_tag_lists)):
        ws.append(wb.create_sheet(title=book_tag_lists[i].decode())) #utf8->unicode
    for i in range(len(book_tag_lists)): 
        ws[i].append(['序号','书名','评分','评价人数','作者','出版社'])
        count=1
        for bl in book_lists[i]:
            ws[i].append([count,bl[0],float(bl[1]),int(bl[2]),bl[3],bl[4]])
            count+=1

    # 获取 存储路径
    f =  contentClass()
    save_path= f.getContentForDouBan(BOOk_TYPE)
    print('save path:',save_path)

    for i in range(len(book_tag_lists)):
        save_path+=('-'+book_tag_lists[i].decode())
    save_path+='.xlsx'
    wb.save(save_path)
    print('save finish,at',save_path)


def create_db():
    do_init_db('douban')
    print('db finish')

